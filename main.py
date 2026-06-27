import os
import time
import hmac
import hashlib
import json
import uuid
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from dotenv import load_dotenv
import razorpay
import pymysql
import pymysql.cursors
from io import BytesIO
from openpyxl import load_workbook

load_dotenv()

# ── Config ─────────────────────────────────────────────────────────────────────
RAZORPAY_KEY_ID     = os.getenv('RAZORPAY_KEY_ID')
RAZORPAY_KEY_SECRET = os.getenv('RAZORPAY_KEY_SECRET')
ALLOWED_ORIGINS     = os.getenv('ALLOWED_ORIGINS', 'http://localhost:8000').split(',')

DB_HOST = os.getenv('DB_HOST', 'localhost')
DB_PORT = int(os.getenv('DB_PORT', '3306'))
DB_NAME = os.getenv('DB_NAME', 'hellobmg')
DB_USER = os.getenv('DB_USER', 'root')
DB_PASS = os.getenv('DB_PASSWORD', '')

# Static image storage
STATIC_DIR = Path(__file__).parent / 'static' / 'images'
STATIC_DIR.mkdir(parents=True, exist_ok=True)
BASE_URL    = os.getenv('BASE_URL', 'http://localhost:2025')

# ── Hardcoded admin user ────────────────────────────────────────────────────────
ADMIN_EMAIL         = 'admin@hellobmg.com'
ADMIN_PASSWORD_HASH = hashlib.sha256('HelloBMG@2025'.encode()).hexdigest()

_active_tokens: set[str] = set()

_security = HTTPBearer()

def require_admin(credentials: HTTPAuthorizationCredentials = Depends(_security)):
    if credentials.credentials not in _active_tokens:
        raise HTTPException(status_code=401, detail='Invalid or expired session. Please sign in.')

# ── App setup ──────────────────────────────────────────────────────────────────
app = FastAPI(title='HelloBMG Backend', version='2.0.0')

app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in ALLOWED_ORIGINS],
    allow_credentials=True,
    allow_methods=['*'],
    allow_headers=['*'],
)

app.mount('/static', StaticFiles(directory=str(Path(__file__).parent / 'static')), name='static')

# ── DB helpers ─────────────────────────────────────────────────────────────────
def get_db() -> pymysql.connections.Connection:
    return pymysql.connect(
        host=DB_HOST, port=DB_PORT,
        user=DB_USER, password=DB_PASS,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=True,
        charset='utf8mb4',
    )

def new_id() -> str:
    return str(uuid.uuid4())

def _ensure_ref(cur, table: str, name: str) -> str:
    """Return id of existing row or insert new one."""
    cur.execute(f'SELECT id FROM {table} WHERE name=%s LIMIT 1', (name,))
    row = cur.fetchone()
    if row:
        return row['id']
    rid = new_id()
    cur.execute(f'INSERT INTO {table} (id, name) VALUES (%s, %s)', (rid, name))
    return rid

# ── Razorpay ───────────────────────────────────────────────────────────────────
if not (RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET):
    print('Warning: Razorpay keys not set. /payments endpoints will fail.')

rz_client: Optional[razorpay.Client] = None
if RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET:
    rz_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))

class CreateOrderRequest(BaseModel):
    amount: int
    currency: str = 'INR'
    notes: Optional[dict] = None

class VerifyRequest(BaseModel):
    razorpay_payment_id: str
    razorpay_order_id: str
    razorpay_signature: str

class LoginRequest(BaseModel):
    email: str
    password: str

@app.post('/admin/login')
async def admin_login(payload: LoginRequest):
    pw_hash = hashlib.sha256(payload.password.encode()).hexdigest()
    if payload.email.strip().lower() != ADMIN_EMAIL or pw_hash != ADMIN_PASSWORD_HASH:
        raise HTTPException(status_code=401, detail='Invalid email or password')
    token = str(uuid.uuid4())
    _active_tokens.add(token)
    return {'token': token, 'email': ADMIN_EMAIL}

@app.post('/admin/logout')
async def admin_logout(credentials: HTTPAuthorizationCredentials = Depends(_security)):
    _active_tokens.discard(credentials.credentials)
    return {'status': 'logged out'}

@app.get('/health')
async def health():
    return {'status': 'ok', 'time': int(time.time())}

@app.post('/payments/create-order')
async def create_order(payload: CreateOrderRequest):
    if rz_client is None:
        raise HTTPException(status_code=500, detail='Razorpay not configured')
    if payload.amount < 100:
        raise HTTPException(status_code=400, detail='Amount must be at least ₹1 (100 paise)')
    try:
        order = rz_client.order.create({
            'amount': payload.amount,
            'currency': payload.currency or 'INR',
            'payment_capture': 1,
            'notes': payload.notes or {},
        })
        return {'key_id': RAZORPAY_KEY_ID, 'order': order}
    except Exception as e:
        raise HTTPException(status_code=500, detail='Order creation failed') from e

@app.post('/payments/verify')
async def verify_signature(payload: VerifyRequest):
    if rz_client is None:
        raise HTTPException(status_code=500, detail='Razorpay not configured')
    body = f'{payload.razorpay_order_id}|{payload.razorpay_payment_id}'
    generated = hmac.new(
        bytes(RAZORPAY_KEY_SECRET, 'utf-8'),
        bytes(body, 'utf-8'),
        hashlib.sha256,
    ).hexdigest()
    if generated != payload.razorpay_signature:
        raise HTTPException(status_code=400, detail='Invalid signature')
    return {'status': 'verified'}

# ── Products (public) ──────────────────────────────────────────────────────────
@app.get('/products')
async def list_products():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("""
                SELECT
                  p.id, p.title, p.price, p.mrp, p.rating, p.stock,
                  p.popularity, p.specs, p.tags, p.stock_status,
                  b.name AS brand_name,
                  c.name AS category_name,
                  pi.local_url AS image_url
                FROM products p
                LEFT JOIN brands b ON b.id = p.brand_id
                LEFT JOIN categories c ON c.id = p.category_id
                LEFT JOIN product_images pi
                  ON pi.product_id = p.id AND pi.is_primary = 1
                WHERE p.is_active = 1
                ORDER BY p.popularity DESC
                LIMIT 200
            """)
            rows = cur.fetchall()
    finally:
        db.close()

    result = []
    for row in rows:
        specs = json.loads(row['specs']) if row['specs'] else {}
        tags  = json.loads(row['tags'])  if row['tags']  else []
        result.append({
            'id':         row['id'],
            'title':      row['title'],
            'brand':      row['brand_name'] or '',
            'model':      specs.get('model', ''),
            'category':   row['category_name'] or 'Other',
            'price':      row['price'],
            'mrp':        row['mrp'],
            'rating':     float(row['rating'] or 0),
            'stock':      row['stock'] or 0,
            'popularity': row['popularity'] or 0,
            'image':      row['image_url'],
            'tags':       tags,
            'specs':      specs,
        })
    return result

# ── Admin: Brands ──────────────────────────────────────────────────────────────
class BrandCreate(BaseModel):
    name: str

@app.get('/admin/brands')
async def list_brands():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute('SELECT id, name FROM brands ORDER BY name')
            return cur.fetchall()
    finally:
        db.close()

@app.post('/admin/brands')
async def create_brand(payload: BrandCreate, _: str = Depends(require_admin)):
    db = get_db()
    try:
        with db.cursor() as cur:
            rid = new_id()
            cur.execute('INSERT INTO brands (id, name) VALUES (%s, %s)', (rid, payload.name.strip()))
            return {'id': rid, 'name': payload.name.strip()}
    except pymysql.IntegrityError:
        raise HTTPException(status_code=400, detail='Brand already exists')
    finally:
        db.close()

# ── Admin: Categories ──────────────────────────────────────────────────────────
class CategoryCreate(BaseModel):
    name: str

@app.get('/admin/categories')
async def list_categories():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute('SELECT id, name FROM categories ORDER BY name')
            return cur.fetchall()
    finally:
        db.close()

@app.post('/admin/categories')
async def create_category(payload: CategoryCreate, _: str = Depends(require_admin)):
    db = get_db()
    try:
        with db.cursor() as cur:
            rid = new_id()
            cur.execute('INSERT INTO categories (id, name) VALUES (%s, %s)', (rid, payload.name.strip()))
            return {'id': rid, 'name': payload.name.strip()}
    except pymysql.IntegrityError:
        raise HTTPException(status_code=400, detail='Category already exists')
    finally:
        db.close()

# ── Admin: Products ────────────────────────────────────────────────────────────
class ProductCreate(BaseModel):
    title: str
    brand_id: Optional[str] = None
    category_id: Optional[str] = None
    price: int
    mrp: Optional[int] = None
    rating: Optional[float] = None
    stock: Optional[int] = 0
    popularity: Optional[int] = 0
    specs: Optional[dict] = None
    tags: Optional[list] = None
    is_active: bool = True
    stock_status: Optional[str] = 'in_stock'

@app.post('/admin/products')
async def create_product(payload: ProductCreate, _: str = Depends(require_admin)):
    db = get_db()
    try:
        with db.cursor() as cur:
            pid = new_id()
            cur.execute("""
                INSERT INTO products
                  (id, title, brand_id, category_id, price, mrp, rating,
                   stock, popularity, specs, tags, is_active, stock_status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                pid, payload.title, payload.brand_id, payload.category_id,
                payload.price, payload.mrp, payload.rating,
                payload.stock or 0, payload.popularity or 0,
                json.dumps(payload.specs or {}),
                json.dumps(payload.tags or []),
                int(payload.is_active),
                payload.stock_status or 'in_stock',
            ))
            return {'id': pid}
    finally:
        db.close()

@app.post('/admin/products/{product_id}/images')
async def upload_product_image(
    product_id: str,
    file: UploadFile = File(...),
    is_primary: bool = Form(True),
    _: str = Depends(require_admin),
):
    content = await file.read()
    ext     = (file.filename or 'image').rsplit('.', 1)[-1].lower() or 'jpg'
    ts      = int(time.time())
    rel     = f'{product_id}/{ts}-primary.{ext}'
    dest    = STATIC_DIR / product_id
    dest.mkdir(parents=True, exist_ok=True)
    (dest / f'{ts}-primary.{ext}').write_bytes(content)
    local_url = f'{BASE_URL}/static/images/{rel}'

    db = get_db()
    try:
        with db.cursor() as cur:
            if is_primary:
                cur.execute(
                    'UPDATE product_images SET is_primary=0 WHERE product_id=%s', (product_id,)
                )
            img_id = new_id()
            cur.execute("""
                INSERT INTO product_images (id, product_id, storage_path, local_url, is_primary, sort_order)
                VALUES (%s,%s,%s,%s,%s,0)
            """, (img_id, product_id, rel, local_url, int(is_primary)))
        return {'path': rel, 'public_url': local_url}
    finally:
        db.close()

class StockUpdate(BaseModel):
    stock: int
    is_active: Optional[bool] = None
    stock_status: Optional[str] = None

@app.patch('/admin/products/{product_id}/stock')
async def update_stock(product_id: str, payload: StockUpdate, _: str = Depends(require_admin)):
    db = get_db()
    try:
        with db.cursor() as cur:
            fields = ['stock=%s']
            vals   = [payload.stock]
            if payload.is_active is not None:
                fields.append('is_active=%s'); vals.append(int(payload.is_active))
            if payload.stock_status is not None:
                fields.append('stock_status=%s'); vals.append(payload.stock_status)
            vals.append(product_id)
            cur.execute(f'UPDATE products SET {", ".join(fields)} WHERE id=%s', vals)
        return {'status': 'ok'}
    finally:
        db.close()

# ── Admin: Bulk XLSX ───────────────────────────────────────────────────────────
@app.post('/admin/products/bulk-xlsx')
async def bulk_upload_xlsx(file: UploadFile = File(...), _: str = Depends(require_admin)):
    content = await file.read()
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail='Empty XLSX file')

    header = [str(h).strip().lower() if h is not None else '' for h in rows[0]]

    def col(name):
        try: return header.index(name)
        except ValueError: return -1

    idx = {k: col(k) for k in [
        'title', 'brand', 'category', 'price', 'mrp', 'rating',
        'stock', 'tags', 'image_url', 'specs_json',
        'storage', 'ram', 'display', 'battery', 'camera', 'warranty', 'stock_status',
    ]}
    for rc in ['title', 'brand', 'category', 'price']:
        if idx[rc] == -1:
            raise HTTPException(status_code=400, detail=f'Missing required column: {rc}')

    created = updated = images_uploaded = 0
    db = get_db()
    try:
        with db.cursor() as cur:
            for r in rows[1:]:
                if not r: continue
                def cell(k, default=None):
                    i = idx.get(k, -1)
                    return r[i] if i >= 0 and i < len(r) and r[i] is not None else default

                title = (cell('title') or '').strip()
                if not title: continue

                brand_name    = (cell('brand') or '').strip()
                category_name = (cell('category') or '').strip()
                price         = int(cell('price') or 0)
                mrp           = int(cell('mrp')) if cell('mrp') is not None else None
                rating        = float(cell('rating')) if cell('rating') is not None else None
                stock         = int(cell('stock') or 0)
                tags_raw      = str(cell('tags') or '')
                tags          = [t.strip() for t in tags_raw.split(',') if t.strip()]
                stock_status  = (str(cell('stock_status') or '') or ('in_stock' if stock > 0 else 'out_of_stock')).strip().lower()

                specs = {}
                if cell('specs_json'):
                    try: specs = json.loads(cell('specs_json'))
                    except Exception: pass
                for key in ['storage', 'ram', 'display', 'battery', 'camera', 'warranty']:
                    if cell(key): specs[key] = str(cell(key))

                brand_id    = _ensure_ref(cur, 'brands', brand_name) if brand_name else None
                category_id = _ensure_ref(cur, 'categories', category_name) if category_name else None

                cur.execute('SELECT id FROM products WHERE title=%s LIMIT 1', (title,))
                existing = cur.fetchone()

                if existing:
                    pid = existing['id']
                    cur.execute("""
                        UPDATE products
                        SET brand_id=%s, category_id=%s, price=%s, mrp=%s, rating=%s,
                            stock=%s, specs=%s, tags=%s, is_active=1, stock_status=%s
                        WHERE id=%s
                    """, (brand_id, category_id, price, mrp, rating, stock,
                          json.dumps(specs), json.dumps(tags), stock_status, pid))
                    updated += 1
                else:
                    pid = new_id()
                    cur.execute("""
                        INSERT INTO products
                          (id, title, brand_id, category_id, price, mrp, rating,
                           stock, popularity, specs, tags, is_active, stock_status)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,0,%s,%s,1,%s)
                    """, (pid, title, brand_id, category_id, price, mrp, rating,
                          stock, json.dumps(specs), json.dumps(tags), stock_status))
                    created += 1

                image_url = cell('image_url', '').strip()
                if image_url:
                    try:
                        import requests as req_lib
                        resp = req_lib.get(image_url, timeout=15)
                        if resp.status_code == 200:
                            ct  = resp.headers.get('content-type', '')
                            ext = 'png' if 'png' in ct else 'jpg'
                            ts  = int(time.time())
                            rel = f'{pid}/{ts}-primary.{ext}'
                            dest = STATIC_DIR / pid
                            dest.mkdir(parents=True, exist_ok=True)
                            (dest / f'{ts}-primary.{ext}').write_bytes(resp.content)
                            local_url = f'{BASE_URL}/static/images/{rel}'
                            cur.execute('UPDATE product_images SET is_primary=0 WHERE product_id=%s', (pid,))
                            img_id = new_id()
                            cur.execute("""
                                INSERT INTO product_images (id, product_id, storage_path, local_url, is_primary, sort_order)
                                VALUES (%s,%s,%s,%s,1,0)
                            """, (img_id, pid, rel, local_url))
                            images_uploaded += 1
                    except Exception:
                        pass

        return {'created': created, 'updated': updated, 'images_uploaded': images_uploaded}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'Bulk upload failed: {e}')
    finally:
        db.close()

# Run: uvicorn main:app --host 0.0.0.0 --port 2025 --reload
